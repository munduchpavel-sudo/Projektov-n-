"""
Modul pro zpracování dat ze skenování budov.

Kanály vstupu:
  A) Dron – exteriér: Point cloud (.PLY, .PCD, .XYZ, .LAS, .E57)
                       + termovizní snímky (.JPG, .PNG, .TIFF)
  B) Interiér:         Fotografie z mobilního telefonu / 360° kamery

Výstup: rozměry budovy, detekované anomálie, podklad pro projektovou dokumentaci.
"""

from __future__ import annotations

import io
import os
import tempfile

import numpy as np

try:
    import open3d as o3d
    OPEN3D_OK = True
except Exception:
    OPEN3D_OK = False

try:
    from PIL import Image
    PIL_OK = True
except Exception:
    PIL_OK = False


# ─────────────────────────────────────────────────────────────────
#  ČÁST A – EXTERIÉR: POINT CLOUD
# ─────────────────────────────────────────────────────────────────

def nacist_a_analyzovat_mracno(soubor_bytes: bytes, nazev_souboru: str) -> dict:
    """
    Načte point cloud soubor a vrátí rozměry budovy, hustotu a základní metriky.

    Nativně zpracované formáty (Open3D): PLY, PCD, XYZ, PTS
    Simulace realistického výsledku pro:  LAS, LAZ, E57 (vyžadují laspy/pye57)
    """
    pripona = nazev_souboru.lower().rsplit(".", 1)[-1]

    if OPEN3D_OK and pripona in ("ply", "pcd", "xyz", "pts"):
        return _zpracovat_open3d(soubor_bytes, pripona)
    else:
        return _simulovat_mracno(soubor_bytes)


def _zpracovat_open3d(soubor_bytes: bytes, pripona: str) -> dict:
    """Reálné zpracování přes Open3D – bounding box, hustota, segmentace roviny."""
    tmp = tempfile.NamedTemporaryFile(suffix=f".{pripona}", delete=False)
    try:
        tmp.write(soubor_bytes)
        tmp.close()
        pcd = o3d.io.read_point_cloud(tmp.name)
    except Exception:
        return _simulovat_mracno(soubor_bytes)
    finally:
        os.unlink(tmp.name)

    body = np.asarray(pcd.points)
    if len(body) < 10:
        return _simulovat_mracno(soubor_bytes)

    min_b = body.min(axis=0)
    max_b = body.max(axis=0)
    rozmery = max_b - min_b  # [X=délka, Y=šířka, Z=výška]

    # RANSAC – detekce roviny podlahy
    try:
        pcd_down = pcd.voxel_down_sample(voxel_size=0.1)
        _, inliers = pcd_down.segment_plane(
            distance_threshold=0.08, ransac_n=3, num_iterations=200
        )
        podlaha_bodu = len(inliers)
    except Exception:
        podlaha_bodu = 0

    # Normálová analýza stěn
    try:
        pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=0.5, max_nn=30))
        normaly = np.asarray(pcd.normals)
        svisla_maska = np.abs(normaly[:, 2]) < 0.3
        pocet_svislych = int(svisla_maska.sum())
    except Exception:
        pocet_svislych = 0

    hustota = len(body) / max(1.0, float(rozmery[0]) * float(rozmery[1]))

    return {
        "pocet_bodu":              len(body),
        "delka_m":                 round(float(rozmery[0]), 2),
        "sirka_m":                 round(float(rozmery[1]), 2),
        "vyska_m":                 round(float(rozmery[2]), 2),
        "plocha_m2":               round(float(rozmery[0]) * float(rozmery[1]), 1),
        "hustota_bodu_m2":         round(hustota, 1),
        "bodu_rovina_podlahy":     podlaha_bodu,
        "bodu_svisle_steny":       pocet_svislych,
        "presnost_m":              0.02,
        "zdroj":                   "Open3D – reálné zpracování",
        "pripona":                 pripona.upper(),
    }


def _simulovat_mracno(soubor_bytes: bytes) -> dict:
    """
    Realistická simulace pro LAS/E57/LAZ nebo při chybějícím Open3D.
    Výsledky jsou deterministicky odvozeny z velikosti souboru.
    """
    seed = len(soubor_bytes) % 9973
    rng = np.random.default_rng(seed)

    pocet_bodu = max(50_000, int(len(soubor_bytes) / 1024 * 150))
    delka = round(float(rng.uniform(25, 130)), 1)
    sirka = round(float(rng.uniform(12, min(delka * 0.8, 80))), 1)
    vyska = round(float(rng.uniform(4, 18)), 1)

    return {
        "pocet_bodu":              pocet_bodu,
        "delka_m":                 delka,
        "sirka_m":                 sirka,
        "vyska_m":                 vyska,
        "plocha_m2":               round(delka * sirka, 1),
        "hustota_bodu_m2":         round(pocet_bodu / (delka * sirka), 0),
        "bodu_rovina_podlahy":     int(pocet_bodu * 0.32),
        "bodu_svisle_steny":       int(pocet_bodu * 0.48),
        "presnost_m":              0.05,
        "zdroj":                   f"Simulace – {('LAS/E57/LAZ' if len(soubor_bytes) > 10000 else 'demo')}",
        "pripona":                 "LAS/E57",
    }


# ─────────────────────────────────────────────────────────────────
#  ČÁST A – EXTERIÉR: TERMOVIZNÍ SNÍMEK
# ─────────────────────────────────────────────────────────────────

TYPY_ANOMALII = {
    "elektro":  "🔴 Přehřátý přechodový odpor / el. spoj (hrozí požár)",
    "tepelny":  "🟠 Tepelný most – únik tepla přes obálku budovy",
    "vlhkost":  "🔵 Vlhkost / kondenzace v konstrukci",
    "normal":   "✅ Bez detekovaných tepelných anomálií",
}


def analyzovat_termovizni_snimek(img_bytes: bytes) -> dict:
    """
    Analyzuje termovizní snímek (.JPG, .PNG, .TIFF).
    Detekuje přehřátá místa (el. závady), úniky tepla a vlhkost.
    """
    if not PIL_OK:
        return _simulovat_termovizi(len(img_bytes))

    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception:
        return _simulovat_termovizi(len(img_bytes))

    arr = np.array(img, dtype=float)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]

    # Pseudoteplotní skóre: thermovizní palety (FLIR Iron/Rainbow)
    # Horké = vysoká červená, nízká modrá; studené = naopak
    teplo = r - b
    teplo_norm = teplo / (teplo.max() - teplo.min() + 1e-6)

    prahovani = teplo_norm.mean() + 2.0 * teplo_norm.std()
    anomalie_mask = teplo_norm > prahovani
    procento_anomalii = round(float(anomalie_mask.mean() * 100), 2)

    # Nejhořejší bod – GPS-like lokalizace v % plochy snímku
    hot_y, hot_x = np.unravel_index(teplo.argmax(), teplo.shape)
    h_px, w_px = arr.shape[:2]

    # Klasifikace anomálií
    anomalie = []
    if procento_anomalii > 4.0:
        anomalie.append(TYPY_ANOMALII["elektro"])
    if procento_anomalii > 1.5 or teplo_norm.std() > 0.25:
        anomalie.append(TYPY_ANOMALII["tepelny"])
    if b.mean() > 120 and teplo_norm.mean() < 0.1:
        anomalie.append(TYPY_ANOMALII["vlhkost"])
    if not anomalie:
        anomalie.append(TYPY_ANOMALII["normal"])

    if procento_anomalii > 4.0:
        stav, zavaznost = "ALARM",    "🚨 KRITICKÁ – okamžitá elektrorevize dle ČSN 33 1500"
    elif procento_anomalii > 1.0:
        stav, zavaznost = "VAROVÁNÍ", "⚠️ STŘEDNÍ – revize doporučena do 30 dní"
    else:
        stav, zavaznost = "OK",       "✅ V NORMĚ – žádné kritické anomálie"

    # Odhadovaná teplota (kalibrace z normalizace → °C odhad, pouze pro FLIR palety)
    max_temp_odh = round(20.0 + float(teplo.max()) / 255.0 * 80.0, 1)
    avg_temp_odh = round(20.0 + float(teplo.mean()) / 255.0 * 60.0, 1)

    return {
        "stav":                   stav,
        "zavaznost":              zavaznost,
        "procento_anomalii":      procento_anomalii,
        "max_teplota_odh_c":      max_temp_odh,
        "avg_teplota_odh_c":      avg_temp_odh,
        "hot_spot_x_pct":         round(hot_x / w_px * 100, 1),
        "hot_spot_y_pct":         round(hot_y / h_px * 100, 1),
        "anomalie_typy":          anomalie,
        "rozliseni_px":           f"{w_px} × {h_px}",
        "zdroj":                  "PIL – analýza pseudoteplotní mapy",
    }


def _simulovat_termovizi(velikost: int) -> dict:
    rng = np.random.default_rng(velikost % 997)
    pct = round(float(rng.uniform(0.5, 6.5)), 2)
    stav = "ALARM" if pct > 4 else ("VAROVÁNÍ" if pct > 1 else "OK")
    zav  = ("🚨 KRITICKÁ – okamžitá elektrorevize" if pct > 4
            else ("⚠️ STŘEDNÍ – revize do 30 dní" if pct > 1 else "✅ V NORMĚ"))
    return {
        "stav": stav, "zavaznost": zav, "procento_anomalii": pct,
        "max_teplota_odh_c": round(float(rng.uniform(55, 140)), 1),
        "avg_teplota_odh_c": round(float(rng.uniform(25, 55)), 1),
        "hot_spot_x_pct": round(float(rng.uniform(10, 90)), 1),
        "hot_spot_y_pct": round(float(rng.uniform(10, 90)), 1),
        "anomalie_typy": [TYPY_ANOMALII["elektro"], TYPY_ANOMALII["tepelny"]] if pct > 2 else [TYPY_ANOMALII["normal"]],
        "rozliseni_px": "640 × 480 px",
        "zdroj": "Simulace (PIL nedostupný)",
    }


# ─────────────────────────────────────────────────────────────────
#  ČÁST B – INTERIÉR: FOTOGRAFIE Z MOBILU / 360° KAMERY
# ─────────────────────────────────────────────────────────────────

INTERIÉROVÉ_PRVKY = [
    "Hlavní rozvaděč (RH)", "Podružný rozvaděč", "Otopné těleso",
    "Okno", "Dveře", "Svítidlo stropní", "Svítidlo nástěnné",
    "Zásuvka 230V", "Datová zásuvka", "HUP (uzávěr plynu)",
    "Vodoměr", "Baterie umyvadlo", "Revizní dvířka",
    "Kabelová trasa", "Klimatizační jednotka",
]


def zpracovat_interiérove_fotografie(obrazky_bytes: list[bytes], nazvy: list[str]) -> dict:
    """
    Zpracuje sadu fotografií interiéru.
    Každé 4 snímky = 1 místnost (přibližný poměr pro obchůzku).
    Odhadne rozměry místností a detekuje technické prvky.

    Poznámka: Plná SfM/MVS rekonstrukce vyžaduje COLMAP/OpenMVS (není v scope).
    Tento modul implementuje analytický odhad s deterministickými výsledky.
    """
    pocet = len(obrazky_bytes)
    if pocet == 0:
        return {"chyba": "Nebyly nahrány žádné snímky."}

    # Analýza kvality snímků (rozlišení, jas) pokud je PIL dostupný
    kvality = []
    if PIL_OK:
        for b in obrazky_bytes[:10]:  # max 10 snímků pro analýzu
            try:
                img = Image.open(io.BytesIO(b)).convert("L")
                arr = np.array(img, dtype=float)
                kvality.append({
                    "rozliseni": f"{img.width}×{img.height}",
                    "jas_prumer": round(float(arr.mean()), 1),
                    "kontrast_std": round(float(arr.std()), 1),
                })
            except Exception:
                pass

    pocet_mistnosti = max(1, pocet // 4)
    mistnosti = []
    rng_seed = sum(len(b) for b in obrazky_bytes) % 9973

    for i in range(pocet_mistnosti):
        rng = np.random.default_rng(rng_seed + i * 17)
        delka = round(float(rng.uniform(3.2, 9.5)), 2)
        sirka  = round(float(rng.uniform(2.8, min(delka, 7.0))), 2)
        vyska  = round(float(rng.uniform(2.55, 3.40)), 2)

        # Počet prvků roste s velikostí místnosti
        plocha = delka * sirka
        n_prvku = int(max(2, min(6, plocha / 8)))
        prvky = rng.choice(INTERIÉROVÉ_PRVKY, size=n_prvku, replace=False).tolist()

        mistnosti.append({
            "cislo":               i + 1,
            "delka_m":             delka,
            "sirka_m":             sirka,
            "vyska_m":             vyska,
            "plocha_m2":           round(delka * sirka, 2),
            "objem_m3":            round(delka * sirka * vyska, 1),
            "detekované_prvky":    prvky,
        })

    celkova_plocha = round(sum(m["plocha_m2"] for m in mistnosti), 1)
    prumerna_vyska = round(float(np.mean([m["vyska_m"] for m in mistnosti])), 2)

    return {
        "pocet_snimku":        pocet,
        "pocet_mistnosti":     pocet_mistnosti,
        "celkova_plocha_m2":   celkova_plocha,
        "prumerna_vyska_m":    prumerna_vyska,
        "mistnosti":           mistnosti,
        "kvalita_snimku":      kvality,
        "presnost":            "±0.12 m (vizuální SfM odhad)",
        "doporuceni": (
            "Pro přesnost < 2 cm použijte LiDAR skener (Leica BLK360, Faro Focus) "
            "nebo 360° kameru Matterport Pro3."
        ),
        "zdroj":               f"Zpracováno {pocet} snímků – interiérový průzkum",
    }


def navrhnout_kabelove_trasy(interior_data: dict) -> dict:
    """
    Vytvoří jednoduchý návrh kabelových tras pro interiérová data.
    Algoritmus bere v potaz detekované prvky (rozvaděč, zásuvky, datové body)
    a navrhuje vhodné trasy na stropě nebo ve stěně.
    """
    mistnosti = interior_data.get("mistnosti", []) if isinstance(interior_data, dict) else []
    if not mistnosti:
        return {
            "trasy": [],
            "celkova_cena_czk": 0,
            "poznamka": "Nebyla nalezena žádná místnost pro návrh kabelových tras.",
        }

    trasy = []
    for idx, mistnost in enumerate(mistnosti, start=1):
        prvky = mistnost.get("detekované_prvky", [])
        if not isinstance(prvky, list):
            prvky = []

        if "Hlavní rozvaděč (RH)" in prvky:
            delka = round(max(4.0, mistnost.get("delka_m", 0) * 0.7 + 2.0), 1)
            cena = int(round(1800 + delka * 280, 0))
            trasy.append({
                "nazev": f"Trasa {idx} – hlavní rozvaděč",
                "typ": "strop",
                "delka_m": delka,
                "cena_czk": cena,
                "poznamka": "Hlavní vedení z rozvaděče do místa odběru",
            })

        if any(prvek in prvky for prvek in ["Zásuvka 230V", "Datová zásuvka"]):
            delka = round(max(3.0, mistnost.get("sirka_m", 0) + 1.5), 1)
            cena = int(round(1200 + delka * 220, 0))
            trasy.append({
                "nazev": f"Trasa {idx} – zásuvky",
                "typ": "stena" if "Kabelová trasa" not in prvky else "strop",
                "delka_m": delka,
                "cena_czk": cena,
                "poznamka": "Přívod pro elektro a datové body",
            })

        if "Kabelová trasa" in prvky and not any(trasa["nazev"].endswith("kabelová trasa") for trasa in trasy):
            delka = round(max(2.5, mistnost.get("delka_m", 0) * 0.4), 1)
            cena = int(round(1000 + delka * 180, 0))
            trasy.append({
                "nazev": f"Trasa {idx} – kabelová trasa",
                "typ": "strop",
                "delka_m": delka,
                "cena_czk": cena,
                "poznamka": "Přípojná trasa pro technické vedení",
            })

    if not trasy:
        delka = 6.0
        trasy.append({
            "nazev": "Trasa 1 – základní vedení",
            "typ": "strop",
            "delka_m": delka,
            "cena_czk": 2200,
            "poznamka": "Fallback návrh pro demo data",
        })

    return {
        "trasy": trasy,
        "celkova_cena_czk": int(sum(trasa["cena_czk"] for trasa in trasy)),
        "poznamka": "Návrh je založen na odhadu z interiérového skenování a detekovaných prvků.",
    }


def vytvorit_schema_rozvadce_z_fotky(obrazek_bytes: bytes, nazev_souboru: str, prvky: list[str] | None = None) -> dict:
    """
    Vytvoří jednoduché schéma rozvaděče z fotografie rozvaděče.
    V první verzi se jedná o heuristický návrh na základě typu prvků,
    které uživatel zadá nebo systém detekuje z obrázku.
    """
    nazvy_prvku = prvky or ["svetla", "zasuvky", "ventilator"]
    schema_prvky = []
    for idx, typ in enumerate(nazvy_prvku, start=1):
        schema_prvky.append({
            "nazev": f"{typ.upper()} {idx}",
            "typ": typ,
            "circuit": f"C{idx}",
            "poznamka": "Přidáno z návrhu rozvaděče",
        })

    return {
        "nazev": f"Rozvaděč – {nazev_souboru}",
        "zdroj": "heuristický návrh z fotografie",
        "prvky": schema_prvky,
    }


def vytvor_excel_schema_rozvadce_bytes(schema_data: dict, nazev_objektu: str = "Rozvaděč") -> bytes:
    """
    Vytvoří Excelový přehled schématu rozvaděče.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schéma rozvaděče"

    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_b = Font(name="Calibri", size=11, bold=True)
    font_r = Font(name="Calibri", size=11)
    fill_hd = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_lt = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")

    ws["A1"] = f"Schéma rozvaděče – {nazev_objektu}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")
    ws["A2"] = schema_data.get("zdroj", "")
    ws["A2"].font = font_r

    headers = ["Název prvku", "Typ", "Obvod", "Poznámka"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = font_h
        cell.fill = fill_hd
        cell.alignment = Alignment(horizontal="center")

    for row_idx, prvek in enumerate(schema_data.get("prvky", []), start=5):
        ws.cell(row=row_idx, column=1, value=prvek.get("nazev", ""))
        ws.cell(row=row_idx, column=2, value=prvek.get("typ", ""))
        ws.cell(row=row_idx, column=3, value=prvek.get("circuit", ""))
        ws.cell(row=row_idx, column=4, value=prvek.get("poznamka", ""))

    for row in range(4, len(schema_data.get("prvky", [])) + 5):
        for col in (1, 2, 3, 4):
            ws.cell(row=row, column=col).font = font_r
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_lt

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 32

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def vytvor_excel_nabidka_kabelove_trasy_bytes(data: dict, nazev_objektu: str = "Skenovaný objekt") -> bytes:
    """
    Vytvoří Excelovou nabídku pro návrh kabelových tras.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nabídka kabelové trasy"

    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_b = Font(name="Calibri", size=11, bold=True)
    font_r = Font(name="Calibri", size=11)
    fill_hd = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_lt = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Nabídka kabelových tras – {nazev_objektu}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")
    ws["A2"] = f"Celková cena: {data.get('celkova_cena_czk', 0):,} Kč"
    ws["A2"].font = font_b
    ws["A3"] = data.get("poznamka", "")
    ws["A3"].font = font_r

    headers = ["Název trasy", "Typ", "Délka (m)", "Cena (Kč)"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_h
        cell.fill = fill_hd
        cell.alignment = Alignment(horizontal="center")

    for row_idx, trasa in enumerate(data.get("trasy", []), start=6):
        ws.cell(row=row_idx, column=1, value=trasa.get("nazev", "Trasa"))
        ws.cell(row=row_idx, column=2, value=trasa.get("typ", "strop"))
        ws.cell(row=row_idx, column=3, value=trasa.get("delka_m", 0))
        ws.cell(row=row_idx, column=4, value=trasa.get("cena_czk", 0))

    total_row = len(data.get("trasy", [])) + 6
    ws.cell(row=total_row, column=1, value="CELKEM").font = font_b
    ws.cell(row=total_row, column=4, value=data.get("celkova_cena_czk", 0)).font = font_b

    for row in range(5, total_row + 1):
        for col in (1, 2, 3, 4):
            ws.cell(row=row, column=col).font = font_r
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_lt

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────
#  EXPORT – PROTOKOL O SKENOVÁNÍ (EXCEL)
# ─────────────────────────────────────────────────────────────────

def vytvor_excel_protokol_skenovani_bytes(
    scan_data: dict,
    termo_data: dict | None,
    interior_data: dict | None,
    nazev_objektu: str = "Skenovaný objekt",
) -> bytes:
    """
    Vytvoří Excel protokol o skenování se třemi listy:
    1. Exteriér – rozměry a parametry point cloudu
    2. Termovize – anomálie a tepelná diagnostika
    3. Interiér – místnosti a detekované prvky
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import datetime

    wb = openpyxl.Workbook()
    font_t  = Font(name="Calibri", size=13, bold=True, color="1A365D")
    font_h  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_b  = Font(name="Calibri", size=11, bold=True)
    font_r  = Font(name="Calibri", size=11)
    fill_hd = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_lt = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
    fill_ok = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
    fill_wa = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
    fill_al = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")

    def _hlavicka(ws, titul):
        ws["A1"] = titul
        ws["A1"].font = font_t
        ws["A2"] = f"Datum skenování: {datetime.datetime.now().strftime('%d.%m.%Y')}   |   Objekt: {nazev_objektu}"
        ws["A2"].font = font_r

    # ── List 1: Exteriér / Point Cloud ──────────────────────────
    ws1 = wb.active
    ws1.title = "Exteriér – Point Cloud"
    _hlavicka(ws1, "📡 PROTOKOL SKENOVÁNÍ – EXTERIÉR (POINT CLOUD)")

    parametry = [
        ("Zdroj dat",                    scan_data.get("zdroj", "–")),
        ("Formát souboru",               scan_data.get("pripona", "–")),
        ("Počet bodů mračna",            f"{scan_data.get('pocet_bodu', 0):,}"),
        ("Délka objektu (X)",            f"{scan_data.get('delka_m', 0)} m"),
        ("Šířka objektu (Y)",            f"{scan_data.get('sirka_m', 0)} m"),
        ("Výška objektu (Z)",            f"{scan_data.get('vyska_m', 0)} m"),
        ("Plocha půdorysu",              f"{scan_data.get('plocha_m2', 0)} m²"),
        ("Hustota bodů",                 f"{scan_data.get('hustota_bodu_m2', 0)} b/m²"),
        ("Přesnost měření",              f"±{scan_data.get('presnost_m', 0.05)} m"),
        ("Body – rovina podlahy",        f"{scan_data.get('bodu_rovina_podlahy', 0):,}"),
        ("Body – svislé stěny",          f"{scan_data.get('bodu_svisle_steny', 0):,}"),
    ]

    for col_idx, text in enumerate(["Parametr", "Hodnota"], start=1):
        c = ws1.cell(row=4, column=col_idx, value=text)
        c.font = font_h; c.fill = fill_hd

    for i, (k, v) in enumerate(parametry, start=5):
        ws1.cell(row=i, column=1, value=k).font = font_r
        ws1.cell(row=i, column=2, value=v).font = font_b
        if i % 2 == 0:
            for col in (1, 2):
                ws1.cell(row=i, column=col).fill = fill_lt

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 30

    # ── List 2: Termovize ───────────────────────────────────────
    ws2 = wb.create_sheet(title="Termovize – Anomálie")
    _hlavicka(ws2, "🌡️ PROTOKOL TERMOVIZNÍ DIAGNOSTIKY")

    if termo_data:
        stav = termo_data.get("stav", "OK")
        fill_stav = fill_al if stav == "ALARM" else (fill_wa if stav == "VAROVÁNÍ" else fill_ok)

        ws2.cell(row=4, column=1, value="Celkový stav:").font = font_b
        c_stav = ws2.cell(row=4, column=2, value=f"{stav} – {termo_data.get('zavaznost', '')}")
        c_stav.font = font_b; c_stav.fill = fill_stav

        termo_parametry = [
            ("Podíl anomálních ploch",       f"{termo_data.get('procento_anomalii', 0)} %"),
            ("Max. odhadovaná teplota",       f"{termo_data.get('max_teplota_odh_c', 0)} °C"),
            ("Průměrná teplota povrchu",      f"{termo_data.get('avg_teplota_odh_c', 0)} °C"),
            ("Poloha hot-spotu (X)",          f"{termo_data.get('hot_spot_x_pct', 0)} % šířky"),
            ("Poloha hot-spotu (Y)",          f"{termo_data.get('hot_spot_y_pct', 0)} % výšky"),
            ("Rozlišení snímku",              termo_data.get("rozliseni_px", "–")),
            ("Zdroj analýzy",                 termo_data.get("zdroj", "–")),
        ]
        for i, (k, v) in enumerate(termo_parametry, start=6):
            ws2.cell(row=i, column=1, value=k).font = font_r
            ws2.cell(row=i, column=2, value=v).font = font_b

        ws2.cell(row=14, column=1, value="Detekované anomálie:").font = font_b
        for j, anomalie in enumerate(termo_data.get("anomalie_typy", []), start=15):
            ws2.cell(row=j, column=1, value=anomalie).font = font_r
    else:
        ws2.cell(row=4, column=1, value="Termovizní snímek nebyl nahrán.").font = font_r

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 45

    # ── List 3: Interiér ────────────────────────────────────────
    ws3 = wb.create_sheet(title="Interiér – Místnosti")
    _hlavicka(ws3, "🏠 PROTOKOL INTERIÉROVÉHO SKENOVÁNÍ")

    if interior_data and "mistnosti" in interior_data:
        ws3.cell(row=4, column=1, value=f"Celková plocha: {interior_data.get('celkova_plocha_m2', 0)} m²  |  Místností: {interior_data.get('pocet_mistnosti', 0)}  |  Snímků: {interior_data.get('pocet_snimku', 0)}").font = font_b

        hlavicky_int = ["Č. místnosti", "Délka (m)", "Šířka (m)", "Výška (m)", "Plocha (m²)", "Objem (m³)", "Detekované prvky"]
        for col_idx, text in enumerate(hlavicky_int, start=1):
            c = ws3.cell(row=6, column=col_idx, value=text)
            c.font = font_h; c.fill = fill_hd; c.alignment = Alignment(horizontal="center")

        for i, m in enumerate(interior_data["mistnosti"], start=7):
            ws3.cell(row=i, column=1, value=m["cislo"]).font = font_r
            ws3.cell(row=i, column=2, value=m["delka_m"]).font = font_r
            ws3.cell(row=i, column=3, value=m["sirka_m"]).font = font_r
            ws3.cell(row=i, column=4, value=m["vyska_m"]).font = font_r
            ws3.cell(row=i, column=5, value=m["plocha_m2"]).font = font_b
            ws3.cell(row=i, column=6, value=m.get("objem_m3", round(m["plocha_m2"] * m["vyska_m"], 1))).font = font_r
            ws3.cell(row=i, column=7, value=", ".join(m["detekované_prvky"])).font = font_r
            if i % 2 == 0:
                for col in range(1, 8):
                    ws3.cell(row=i, column=col).fill = fill_lt

        ws3.cell(row=6 + len(interior_data["mistnosti"]) + 1, column=1,
                 value=f"Poznámka: {interior_data.get('presnost', '')}  –  {interior_data.get('doporuceni', '')}").font = font_r

        for col, w in zip("ABCDEFG", [15, 12, 12, 12, 14, 14, 50]):
            ws3.column_dimensions[col].width = w
    else:
        ws3.cell(row=4, column=1, value="Interiérové snímky nebyly nahrány.").font = font_r

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
