import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def pridat_energetickou_navratnost_do_excelu(soubor_xlsx, data_energo):
    """
    Načte stávající rozpočtový Excel a přidá do něj detailní kalkulaci
    úspor a investiční návratnosti FVE + baterií.
    """
    wb = openpyxl.load_workbook(soubor_xlsx)

    ws = wb.create_sheet(title="Návratnost FVE a Energetiky")
    ws.sheet_view.showGridLines = True

    font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

    ws["A1"] = "EKONOMICKÁ BILANCE ENERGETICKÉHO PROJEKTU"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")

    hlavicky = ["Ukazatel", "Hodnota", "Jednotka"]
    for col_idx, text in enumerate(hlavicky, start=1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header

    vyrobena_energie_kwh = data_energo["vykon_kwp"] * 1000
    vyuziti_objektem = 0.85
    cena_kwh = 6.0

    rocni_uspora_czk = vyrobena_energie_kwh * vyuziti_objektem * cena_kwh
    odhadovana_investice = (data_energo["vykon_kwp"] * 22000) + 1200000

    radky_dat = [
        ("Instalovaný výkon střešní FVE", data_energo["vykon_kwp"], "kWp"),
        ("Předpokládaná roční výroba", vyrobena_energie_kwh, "kWh"),
        ("Míra přímého využití energie (s BESS)", vyuziti_objektem * 100, "%"),
        ("Roční finanční úspora na nákladech", rocni_uspora_czk, "Kč"),
        ("Celková investice (FVE + BESS + Trafo)", odhadovana_investice, "Kč"),
    ]

    for idx, (ukazatel, hodnota, jednotka) in enumerate(radky_dat, start=4):
        ws.cell(row=idx, column=1, value=ukazatel).font = font_regular
        cell_val = ws.cell(row=idx, column=2, value=hodnota)
        cell_val.font = font_bold
        ws.cell(row=idx, column=3, value=jednotka).font = font_regular
        if jednotka == "Kč":
            cell_val.number_format = '#,##0" Kč"'

    r_roi = 9
    ws.cell(row=r_roi, column=1, value="Prostá návratnost energetické investice:").font = font_bold
    ws.cell(row=r_roi, column=2, value="=B8/B7").font = font_bold
    ws.cell(row=r_roi, column=2).number_format = '0.0" roku"'

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    wb.save(soubor_xlsx)
    print(f"📊 Energetické úspory úspěšně propsány do Excelu: {soubor_xlsx}")


def vytvor_excel_rozpocet_bytes(data_energo):
    """
    Vytvoří nový Excel soubor v paměti s energetickou návratností pro okamžitý export.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Návratnost FVE a Energetiky"
    ws.sheet_view.showGridLines = True

    font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

    ws["A1"] = "EKONOMICKÁ BILANCE ENERGETICKÉHO PROJEKTU"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")

    hlavicky = ["Ukazatel", "Hodnota", "Jednotka"]
    for col_idx, text in enumerate(hlavicky, start=1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header

    vyrobena_energie_kwh = data_energo["vykon_kwp"] * 1000
    vyuziti_objektem = 0.85
    cena_kwh = 6.0

    rocni_uspora_czk = vyrobena_energie_kwh * vyuziti_objektem * cena_kwh
    odhadovana_investice = (data_energo["vykon_kwp"] * 22000) + 1200000

    radky_dat = [
        ("Instalovaný výkon střešní FVE", data_energo["vykon_kwp"], "kWp"),
        ("Předpokládaná roční výroba", vyrobena_energie_kwh, "kWh"),
        ("Míra přímého využití energie (s BESS)", vyuziti_objektem * 100, "%"),
        ("Roční finanční úspora na nákladech", rocni_uspora_czk, "Kč"),
        ("Celková investice (FVE + BESS + Trafo)", odhadovana_investice, "Kč"),
    ]

    for idx, (ukazatel, hodnota, jednotka) in enumerate(radky_dat, start=4):
        ws.cell(row=idx, column=1, value=ukazatel).font = font_regular
        cell_val = ws.cell(row=idx, column=2, value=hodnota)
        cell_val.font = font_bold
        ws.cell(row=idx, column=3, value=jednotka).font = font_regular
        if jednotka == "Kč":
            cell_val.number_format = '#,##0" Kč"'

    r_roi = 9
    ws.cell(row=r_roi, column=1, value="Prostá návratnost energetické investice:").font = font_bold
    ws.cell(row=r_roi, column=2, value="=B8/B7").font = font_bold
    ws.cell(row=r_roi, column=2).number_format = '0.0" roku"'

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def vytvor_excel_rozpocet_s_esg_bytes(data_energo):
    """
    Vytvoří nový Excel soubor v paměti, který obsahuje jak ROI kalkulaci,
    tak ESG/nefinanční reporting o snížení uhlíkové stopy.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI + ESG"
    ws.sheet_view.showGridLines = True

    font_titul = Font(name="Calibri", size=14, bold=True, color="2F855A")
    font_hlavicka = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    fill_hlavicka = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")

    ws["A1"] = "NEFINANČNÍ REPORTING (ESG) – BILANCE EMISÍ CO₂"
    ws["A1"].font = font_titul

    hlavicky = ["Sledovaný ESG ukazatel", "Hodnota", "Jednotka", "Ekvivalent v přírodě"]
    for col_idx, text in enumerate(hlavicky, start=1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_hlavicka
        cell.fill = fill_hlavicka
        cell.alignment = Alignment(horizontal="center")

    rocni_vyroba_mwh = (data_energo["vykon_kwp"] * 1000) / 1000
    emisni_faktor_cr_co2 = 0.42
    rocni_uspora_co2 = rocni_vyroba_mwh * emisni_faktor_cr_co2
    ekvivalent_stromu = (rocni_uspora_co2 * 1000) / 22

    radky_dat = [
        ("Celkový instalovaný bezemisní výkon", data_energo["vykon_kwp"], "kWp", "Fotovoltaická soustava"),
        ("Předpokládaná roční čistá výroba", rocni_vyroba_mwh, "MWh", "Zelená energie"),
        ("Roční snížení uhlíkové stopy (Scope 2)", rocni_uspora_co2, "t CO₂e", f"Ekvivalent {int(ekvivalent_stromu)} vzrostlých stromů"),
        ("Předpokládaná úspora CO₂ za 20 let provozu", rocni_uspora_co2 * 20, "t CO₂e", "Dlouhodobý ekologický přínos"),
    ]

    for idx, (ukazatel, hodnota, jednotka, ekv) in enumerate(radky_dat, start=4):
        ws.cell(row=idx, column=1, value=ukazatel).font = font_regular
        ws.cell(row=idx, column=2, value=hodnota).font = font_bold
        ws.cell(row=idx, column=3, value=jednotka).font = font_regular
        ws.cell(row=idx, column=4, value=ekv).font = font_regular
        if isinstance(hodnota, float):
            ws.cell(row=idx, column=2).number_format = '#,##0.0'

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# Jednotkové ceny stavebních prací (Kč/m² HPP) dle typu budovy
TYPY_BUDOV_ROZPOCET = {
    "Průmyslová hala / Logistika":           {"cena_m2": 9500,   "poznamka": "Ocelový skelet, PIR panely, průmyslová podlaha"},
    "Výrobní závod":                          {"cena_m2": 13000,  "poznamka": "ŽB skelet, technologická podlaha, VZT"},
    "Administrativní budova":                 {"cena_m2": 55000,  "poznamka": "ŽB monolitický skelet, SDK příčky, klimatizace"},
    "Bytový dům":                             {"cena_m2": 48000,  "poznamka": "ŽB monolitický skelet, cihlové příčky"},
    "Rodinný dům":                            {"cena_m2": 42000,  "poznamka": "Zděná konstrukce, dřevěný krov"},
    "Hotel / Penzion":                        {"cena_m2": 62000,  "poznamka": "ŽB skelet, standardní vybavení pokojů"},
    "Nemocnice / Zdravotnické zařízení":      {"cena_m2": 90000,  "poznamka": "Speciální TZB, hygienické standardy ISO"},
    "Škola / Vzdělávací zařízení":            {"cena_m2": 54000,  "poznamka": "ŽB skelet, akustické příčky, VZT"},
    "Obchodní centrum":                       {"cena_m2": 46000,  "poznamka": "Ocelový / ŽB skelet, prosklené fasády"},
    "Loď / Plavidlo (ocelová konstrukce)":    {"cena_m2": 155000, "poznamka": "Ocelový trup, lodní systémy, ISO 9001"},
}

# Procentní rozpad rozpočtu na stavební díly (dle TSKP / třídník)
ROZPOCET_STRUKTURA = [
    ("HSV 1 – Zemní práce a hrubé terénní úpravy",        0.04),
    ("HSV 2 – Základy a spodní stavba",                   0.10),
    ("HSV 3 – Svislé a kompletní konstrukce (stěny, sloupy)", 0.15),
    ("HSV 4 – Vodorovné konstrukce (stropy, schodiště)",  0.10),
    ("HSV 5 – Komunikace a plochy (vnější zpevnění)",     0.04),
    ("HSV 6 – Úpravy povrchů, podlahy",                  0.08),
    ("HSV 8 – Trubní vedení a přípojky IS",               0.03),
    ("PSV 711/712 – Střešní plášť, hydroizolace, krytina", 0.08),
    ("PSV 762 – Fasáda, výplně otvorů (okna, dveře)",     0.11),
    ("PSV 730 – Elektroinstalace silnoproud + LPS",        0.07),
    ("PSV 735 – Slaboproud, EZS, EPS, datové sítě",       0.03),
    ("PSV 720 – Zdravotechnika (voda, kanalizace)",        0.05),
    ("PSV 710 – Vytápění, chlazení a vzduchotechnika (VZT)", 0.08),
    ("Projektová dokumentace a inženýring (DUR + DSP + DSPS)", 0.04),
]


def vytvor_excel_stavebni_rozpocet_bytes(
    typ_budovy: str,
    delka_m: float,
    sirka_m: float,
    vyska_m: float,
    pocet_podlazi: int = 1,
) -> bytes:
    """
    Sestaví podrobný stavební rozpočet dle TSKP (Třídník stavebních konstrukcí a prací).
    Výstupem je Excel se dvěma listy: řádkový rozpočet a rekapitulace.
    """
    konfig = TYPY_BUDOV_ROZPOCET.get(typ_budovy, TYPY_BUDOV_ROZPOCET["Průmyslová hala / Logistika"])

    hpp_m2 = delka_m * sirka_m * pocet_podlazi  # Hrubá podlažní plocha
    obalka_m2 = 2 * (delka_m + sirka_m) * vyska_m * pocet_podlazi  # Obálka stěn
    celkova_cena = hpp_m2 * konfig["cena_m2"]

    wb = openpyxl.Workbook()

    # ── List 1: Řádkový rozpočet ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Rozpočet – stavební díly"

    font_titul  = Font(name="Calibri", size=14, bold=True, color="1A365D")
    font_h      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold   = Font(name="Calibri", size=11, bold=True)
    font_reg    = Font(name="Calibri", size=11)
    fill_modra  = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_svetla = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
    fill_seda   = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    ws["A1"] = f"🏗️ STAVEBNÍ ROZPOČET – {typ_budovy.upper()}"
    ws["A1"].font = font_titul

    # Základní parametry objektu
    params = [
        ("Délka objektu", delka_m, "m"),
        ("Šířka objektu", sirka_m, "m"),
        ("Výška podlaží", vyska_m, "m"),
        ("Počet nadzemních podlaží", pocet_podlazi, "NP"),
        ("Hrubá podlažní plocha (HPP)", hpp_m2, "m²"),
        ("Obálka svislých konstrukcí", obalka_m2, "m²"),
        ("Jednotková cena (Kč/m² HPP)", konfig["cena_m2"], "Kč/m²"),
        ("Konstrukční systém / poznámka", konfig["poznamka"], ""),
    ]
    for i, (nazev, hodnota, jednotka) in enumerate(params, start=3):
        ws.cell(row=i, column=1, value=nazev).font = font_reg
        cell_val = ws.cell(row=i, column=2, value=hodnota)
        cell_val.font = font_bold
        ws.cell(row=i, column=3, value=jednotka).font = font_reg
        if jednotka == "Kč/m²":
            cell_val.number_format = '#,##0" Kč/m²"'
        elif jednotka == "m²":
            cell_val.number_format = '#,##0.0" m²"'

    # Záhlaví tabulky rozpočtu
    r_start = 13
    hlavicky = ["Kód / Díl", "Popis stavebního dílu", "Podíl (%)", "Cena (Kč bez DPH)"]
    for col_idx, text in enumerate(hlavicky, start=1):
        cell = ws.cell(row=r_start, column=col_idx, value=text)
        cell.font = font_h
        cell.fill = fill_modra
        cell.alignment = Alignment(horizontal="center")

    # Řádky rozpočtu
    for i, (popis, podil) in enumerate(ROZPOCET_STRUKTURA, start=r_start + 1):
        cena_dilu = celkova_cena * podil
        kod = popis.split(" – ")[0] if " – " in popis else popis[:8]
        nazev = popis.split(" – ", 1)[1] if " – " in popis else popis
        fill_radku = fill_seda if i % 2 == 0 else None

        ws.cell(row=i, column=1, value=kod).font = font_bold
        ws.cell(row=i, column=2, value=nazev).font = font_reg
        cell_podil = ws.cell(row=i, column=3, value=round(podil * 100, 0))
        cell_podil.number_format = '0"%"'
        cell_podil.font = font_reg
        cell_cena = ws.cell(row=i, column=4, value=cena_dilu)
        cell_cena.number_format = '#,##0" Kč"'
        cell_cena.font = font_reg

        if fill_radku:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = fill_radku

    # Řádek celkem
    r_total = r_start + 1 + len(ROZPOCET_STRUKTURA)
    ws.cell(row=r_total, column=2, value="CELKOVÁ CENA BEZ DPH").font = font_bold
    cell_tot = ws.cell(row=r_total, column=4, value=celkova_cena)
    cell_tot.number_format = '#,##0" Kč"'
    cell_tot.font = font_bold
    cell_tot.fill = fill_svetla
    ws.cell(row=r_total, column=2).fill = fill_svetla

    # DPH 21 %
    dph = celkova_cena * 0.21
    ws.cell(row=r_total + 1, column=2, value="DPH 21 %").font = font_reg
    cell_dph = ws.cell(row=r_total + 1, column=4, value=dph)
    cell_dph.number_format = '#,##0" Kč"'

    ws.cell(row=r_total + 2, column=2, value="CELKOVÁ CENA VČETNĚ DPH").font = font_bold
    cell_sdph = ws.cell(row=r_total + 2, column=4, value=celkova_cena + dph)
    cell_sdph.number_format = '#,##0" Kč"'
    cell_sdph.font = font_bold
    cell_sdph.fill = fill_svetla
    ws.cell(row=r_total + 2, column=2).fill = fill_svetla

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22

    # ── List 2: Rekapitulace ────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Rekapitulace")
    ws2["A1"] = "📋 REKAPITULACE INVESTIČNÍCH NÁKLADŮ"
    ws2["A1"].font = font_titul

    rekap = [
        ("Stavební práce (HSV + PSV) bez projekce", celkova_cena * 0.96, "Kč"),
        ("Projektová dokumentace a inženýring",      celkova_cena * 0.04, "Kč"),
        ("Celková cena stavby (bez DPH)",            celkova_cena,        "Kč"),
        ("DPH 21 %",                                 dph,                 "Kč"),
        ("Celková cena stavby (včetně DPH)",         celkova_cena + dph,  "Kč"),
        ("Jednotková cena na m² HPP (bez DPH)",      konfig["cena_m2"],   "Kč/m²"),
        ("Hrubá podlažní plocha",                    hpp_m2,              "m²"),
    ]
    for i, (nazev, hodnota, jednotka) in enumerate(rekap, start=3):
        ws2.cell(row=i, column=1, value=nazev).font = font_reg
        cell_v = ws2.cell(row=i, column=2, value=hodnota)
        cell_v.font = font_bold
        cell_v.number_format = '#,##0" Kč"' if jednotka == "Kč" else '#,##0" Kč/m²"' if "m²" in jednotka else '#,##0.0" m²"'
        ws2.cell(row=i, column=3, value=jednotka).font = font_reg

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# Konfigurace všech podporovaných typů energetických projektů
KONFIGURACE_ENERGO = {
    "FVE střešní": {
        "ikona": "☀️", "jednotka_vykonu": "kWp",
        "cena_jednotky": 22000, "fixni_naklady": 80000,
        "spec_vyroba_kwh": 1000, "zivotnost": 25,
        "emisni_faktor": 0.42, "legenda_vykonu": "Instalovaný výkon střešní FVE",
    },
    "FVE pozemní": {
        "ikona": "🌞", "jednotka_vykonu": "kWp",
        "cena_jednotky": 18000, "fixni_naklady": 500000,
        "spec_vyroba_kwh": 1050, "zivotnost": 30,
        "emisni_faktor": 0.42, "legenda_vykonu": "Instalovaný výkon pozemní FVE",
    },
    "Větrná elektrárna": {
        "ikona": "💨", "jednotka_vykonu": "kW",
        "cena_jednotky": 55000, "fixni_naklady": 200000,
        "spec_vyroba_kwh": 2200, "zivotnost": 20,
        "emisni_faktor": 0.42, "legenda_vykonu": "Jmenovitý výkon turbíny",
    },
    "Tepelné čerpadlo vzduch/voda": {
        "ikona": "🌡️", "jednotka_vykonu": "kW",
        "cena_jednotky": 18000, "fixni_naklady": 150000,
        "spec_vyroba_kwh": 3500, "zivotnost": 20,
        "emisni_faktor": 0.18, "legenda_vykonu": "Tepelný výkon tepelného čerpadla",
    },
    "BESS – bateriové úložiště": {
        "ikona": "🔋", "jednotka_vykonu": "kWh",
        "cena_jednotky": 8500, "fixni_naklady": 200000,
        "spec_vyroba_kwh": 365, "zivotnost": 15,
        "emisni_faktor": 0.42, "legenda_vykonu": "Kapacita bateriového systému",
    },
    "Nabíjecí stanice EV": {
        "ikona": "🔌", "jednotka_vykonu": "kW",
        "cena_jednotky": 12000, "fixni_naklady": 250000,
        "spec_vyroba_kwh": 2000, "zivotnost": 15,
        "emisni_faktor": 0.42, "legenda_vykonu": "Výkon nabíjecí stanice",
    },
}


def vytvor_excel_energeticky_projekt_bytes(typ_projektu: str, vykon: float, cena_kwh: float = 5.5) -> bytes:
    """
    Vytvoří kompletní dvoulistový Excel s investičním rozpočtem (CAPEX/OPEX/návratnost)
    a ESG CO₂ bilancí pro libovolný typ energetického projektu.
    """
    konfig = KONFIGURACE_ENERGO.get(typ_projektu, KONFIGURACE_ENERGO["FVE střešní"])

    wb = openpyxl.Workbook()

    # --- List 1: Investiční rozpočet ---
    ws1 = wb.active
    ws1.title = "Investiční rozpočet"

    font_titul = Font(name="Calibri", size=14, bold=True, color="1A365D")
    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    fill_modra = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_svetla = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")

    ws1["A1"] = f"{konfig['ikona']} INVESTIČNÍ ROZPOČET – {typ_projektu.upper()}"
    ws1["A1"].font = font_titul
    ws1["A3"] = f"{konfig['legenda_vykonu']}:"
    ws1["B3"] = vykon
    ws1["C3"] = konfig["jednotka_vykonu"]
    for col in ("A3", "B3", "C3"):
        ws1[col].font = font_bold

    ws1["A5"] = "POLOŽKA INVESTIČNÍCH NÁKLADŮ (CAPEX)"
    ws1["A5"].font = font_h
    ws1["A5"].fill = fill_modra
    ws1["B5"] = "Cena (Kč)"
    ws1["B5"].font = font_h
    ws1["B5"].fill = fill_modra

    capex_tech = vykon * konfig["cena_jednotky"]
    capex_projekce = capex_tech * 0.08
    capex_montaz = capex_tech * 0.15
    capex_infra = konfig["fixni_naklady"]
    capex_celkem = capex_tech + capex_projekce + capex_montaz + capex_infra

    polozky_capex = [
        (f"Technologie – {typ_projektu} ({vykon} {konfig['jednotka_vykonu']} × {konfig['cena_jednotky']:,} Kč)", capex_tech),
        ("Projektová dokumentace a inženýring (8 %)", capex_projekce),
        ("Montáž, zprovoznění a výchozí revize (15 %)", capex_montaz),
        ("Elektrická přípojka, trafostanice a rozvaděče (pevná položka)", capex_infra),
        ("CELKOVÁ INVESTICE (CAPEX)", capex_celkem),
    ]

    for i, (popis, hodnota) in enumerate(polozky_capex, start=6):
        je_celkem = "CELKOVÁ" in popis
        ws1.cell(row=i, column=1, value=popis).font = font_bold if je_celkem else font_regular
        cell = ws1.cell(row=i, column=2, value=hodnota)
        cell.number_format = '#,##0" Kč"'
        cell.font = font_bold if je_celkem else font_regular
        if je_celkem:
            cell.fill = fill_svetla
            ws1.cell(row=i, column=1).fill = fill_svetla

    rocni_vyroba = vykon * konfig["spec_vyroba_kwh"]
    rocni_vynosy = rocni_vyroba * cena_kwh
    opex_rocni = capex_celkem * 0.015
    rocni_cisty_vynos = rocni_vynosy - opex_rocni

    ws1["A12"] = "ROČNÍ VÝNOSY / ÚSPORY A PROVOZNÍ NÁKLADY (OPEX)"
    ws1["A12"].font = font_h
    ws1["A12"].fill = fill_modra
    ws1["B12"] = "Hodnota"
    ws1["B12"].font = font_h
    ws1["B12"].fill = fill_modra
    ws1["C12"] = "Jednotka"
    ws1["C12"].font = font_h
    ws1["C12"].fill = fill_modra

    polozky_vynosy = [
        ("Předpokládaná roční výroba / výkon", rocni_vyroba, "kWh"),
        ("Cena energie / výkup", cena_kwh, "Kč/kWh"),
        ("Roční výnosy / úspory (hrubé)", rocni_vynosy, "Kč"),
        ("Roční provozní náklady OPEX (1,5 % CAPEX)", opex_rocni, "Kč"),
        ("Roční čistý výnos po OPEX", rocni_cisty_vynos, "Kč"),
    ]

    for i, (popis, hodnota, jednotka) in enumerate(polozky_vynosy, start=13):
        ws1.cell(row=i, column=1, value=popis).font = font_regular
        cell_val = ws1.cell(row=i, column=2, value=hodnota)
        cell_val.font = font_bold
        ws1.cell(row=i, column=3, value=jednotka).font = font_regular
        if jednotka == "Kč":
            cell_val.number_format = '#,##0" Kč"'
        elif jednotka == "kWh":
            cell_val.number_format = '#,##0" kWh"'

    r_roi = 19
    ws1.cell(row=r_roi, column=1, value="⏱  PROSTÁ NÁVRATNOST INVESTICE (PBP):").font = font_bold
    ws1.cell(row=r_roi, column=2, value=round(capex_celkem / max(rocni_cisty_vynos, 1), 1)).font = font_bold
    ws1.cell(row=r_roi, column=3, value="roku").font = font_bold
    for col in range(1, 4):
        ws1.cell(row=r_roi, column=col).fill = fill_svetla

    ws1.column_dimensions["A"].width = 58
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 14

    # --- List 2: ESG CO₂ bilance ---
    ws2 = wb.create_sheet(title="ESG – CO₂ bilance")
    fill_zelena = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")

    ws2["A1"] = f"🌱 ESG REPORTING – SNÍŽENÍ UHLÍKOVÉ STOPY | {typ_projektu}"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F855A")

    for col_idx, text in enumerate(["ESG ukazatel", "Hodnota", "Jednotka", "Ekvivalent"], start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_zelena
        cell.alignment = Alignment(horizontal="center")

    rocni_co2 = (rocni_vyroba / 1000) * konfig["emisni_faktor"]
    ekvivalent_stromu = int((rocni_co2 * 1000) / 22)

    radky_esg = [
        ("Instalovaný výkon / kapacita", vykon, konfig["jednotka_vykonu"], typ_projektu),
        ("Předpokládaná roční výroba / výkon", rocni_vyroba, "kWh", "Bezemisní / obnovitelná energie"),
        (f"Roční snížení uhlíkové stopy (emisní faktor {konfig['emisni_faktor']} t/MWh)", rocni_co2, "t CO₂e", f"≈ {ekvivalent_stromu} vzrostlých stromů"),
        (f"Kumulativní úspora CO₂ za {konfig['zivotnost']} let životnosti", rocni_co2 * konfig["zivotnost"], "t CO₂e", "Dlouhodobý ekologický přínos"),
    ]

    for idx, (ukazatel, hodnota, jednotka, ekv) in enumerate(radky_esg, start=4):
        ws2.cell(row=idx, column=1, value=ukazatel).font = Font(name="Calibri", size=11)
        ws2.cell(row=idx, column=2, value=hodnota).font = Font(name="Calibri", size=11, bold=True)
        ws2.cell(row=idx, column=3, value=jednotka).font = Font(name="Calibri", size=11)
        ws2.cell(row=idx, column=4, value=ekv).font = Font(name="Calibri", size=11)
        ws2.cell(row=idx, column=2).number_format = '#,##0.0'

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 48

    buffer2 = io.BytesIO()
    wb.save(buffer2)
    buffer2.seek(0)
    return buffer2.getvalue()


def vytvor_excel_schema_rozvadice_a_cenovou_nabidku_bytes(schema_data: dict) -> bytes:
    """
    Vytvori dvoulistovy Excel se schematem rozvadece a cenovou nabidkou.
    Funkce je urcena pro export dat z navrhu elektrocasti projektu.
    """
    wb = openpyxl.Workbook()

    schema_ws = wb.active
    schema_ws.title = "Schema rozvaděče"
    offer_ws = wb.create_sheet(title="Cenová nabídka")

    nazev = schema_data.get("nazev", "Rozvaděč")
    umisteni = schema_data.get("umisteni", "Neuvedeno")
    zakazka = schema_data.get("zakazka", "Neuvedeno")
    prvky = schema_data.get("prvky", [])

    schema_ws["A1"] = f"SCHÉMA ROZVADĚČE - {nazev}"
    schema_ws["A2"] = f"Umístění: {umisteni}"
    schema_ws["A3"] = f"Zakázka: {zakazka}"
    schema_ws["A4"] = "Přehled prvků"

    schema_ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    schema_ws["A4"].font = Font(name="Calibri", size=11, bold=True)

    for idx, prvek in enumerate(prvky, start=5):
        nazev_prvku = prvek.get("nazev", "")
        typ_prvku = prvek.get("typ", "")
        schema_ws.cell(row=idx, column=1, value=f"{nazev_prvku} ({typ_prvku})")

    offer_ws["A1"] = f"CENOVÁ NABÍDKA - {nazev}"
    offer_ws["A1"].font = Font(name="Calibri", size=14, bold=True)

    headers = ["Položka", "Typ", "Množství", "Cena"]
    for col_idx, text in enumerate(headers, start=1):
        cell = offer_ws.cell(row=2, column=col_idx, value=text)
        cell.font = Font(name="Calibri", size=11, bold=True)

    total = 0
    for row_idx, prvek in enumerate(prvky, start=3):
        cena = float(prvek.get("cena", 0) or 0)
        offer_ws.cell(row=row_idx, column=1, value=prvek.get("nazev", ""))
        offer_ws.cell(row=row_idx, column=2, value=prvek.get("typ", ""))
        offer_ws.cell(row=row_idx, column=3, value=1)
        offer_ws.cell(row=row_idx, column=4, value=cena)
        offer_ws.cell(row=row_idx, column=4).number_format = '#,##0" Kč"'
        total += cena

    total_row = max(3, len(prvky) + 3)
    offer_ws.cell(row=total_row, column=3, value="Celkem")
    offer_ws.cell(row=total_row, column=3).font = Font(name="Calibri", size=11, bold=True)
    offer_ws.cell(row=total_row, column=4, value=total)
    offer_ws.cell(row=total_row, column=4).font = Font(name="Calibri", size=11, bold=True)
    offer_ws.cell(row=total_row, column=4).number_format = '#,##0" Kč"'

    schema_ws.column_dimensions["A"].width = 70
    offer_ws.column_dimensions["A"].width = 40
    offer_ws.column_dimensions["B"].width = 18
    offer_ws.column_dimensions["C"].width = 10
    offer_ws.column_dimensions["D"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
