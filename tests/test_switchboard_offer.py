import io

from openpyxl import load_workbook

from excel_energeticka_navratnost import vytvor_excel_schema_rozvadice_a_cenovou_nabidku_bytes


def test_generates_switchboard_schema_and_price_offer_excel():
    data = {
        "nazev": "RH-01",
        "umisteni": "Hala A",
        "zakazka": "SO-01",
        "prvky": [
            {"nazev": "Hlavní jistič 250A", "typ": "3P/4P", "cena": 12500},
            {"nazev": "Výstup 1 - rozvaděč 63A", "typ": "3P", "cena": 4200},
            {"nazev": "Přístrojové rozvaděče", "typ": "komplet", "cena": 8900},
        ],
    }

    workbook_bytes = vytvor_excel_schema_rozvadice_a_cenovou_nabidku_bytes(data)

    assert isinstance(workbook_bytes, bytes)
    assert len(workbook_bytes) > 0

    wb = load_workbook(io.BytesIO(workbook_bytes))
    assert wb.sheetnames == ["Schema rozvaděče", "Cenová nabídka"]

    schema_sheet = wb["Schema rozvaděče"]
    offer_sheet = wb["Cenová nabídka"]

    assert schema_sheet["A1"].value == "SCHÉMA ROZVADĚČE - RH-01"
    assert "Hlavní jistič 250A" in str(schema_sheet["A5"].value)
    assert offer_sheet["A1"].value == "CENOVÁ NABÍDKA - RH-01"
    assert offer_sheet["A3"].value == "Hlavní jistič 250A"
    assert offer_sheet["D3"].value == 12500
